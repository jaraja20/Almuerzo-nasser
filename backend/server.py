from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Header, Depends
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, io, re, uuid, logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Dict, Optional, Any
from datetime import datetime, timezone
import openpyxl
import xlsxwriter

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI()
api_router = APIRouter(prefix="/api")

ADMIN_USER = os.environ.get("ADMIN_USER", "recepcion")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "rec73491654")
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "nasser-admin-secret-2026")

DAYS_ORDER = ["lunes", "martes", "miercoles", "jueves", "viernes", "sabado"]
DAY_LABELS = {
    "lunes": "Lunes", "martes": "Martes", "miercoles": "Miércoles",
    "jueves": "Jueves", "viernes": "Viernes", "sabado": "Sábado",
}

# ---------- Models ----------
class DayMenu(BaseModel):
    day: str  # lunes..sabado
    date: Optional[str] = None  # ISO date string
    breakfast: List[str] = []
    main: List[str] = []
    diet: List[str] = []
    sides: List[str] = []

class Menu(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    provider: str  # mara | sabrositos
    week_start: Optional[str] = None
    week_end: Optional[str] = None
    days: List[DayMenu] = []
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class MenuCreate(BaseModel):
    provider: str
    week_start: Optional[str] = None
    week_end: Optional[str] = None
    days: List[DayMenu] = []

class Selection(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_name: str
    provider: str
    menu_id: str
    week_start: Optional[str] = None
    week_end: Optional[str] = None
    # choices: { "lunes": {"breakfast": "...", "main": "...", "side": "..."}, ... }
    choices: Dict[str, Dict[str, Optional[str]]] = {}
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class SelectionCreate(BaseModel):
    user_name: str
    provider: str
    menu_id: str
    choices: Dict[str, Dict[str, Optional[str]]]

class LoginBody(BaseModel):
    username: str
    password: str

class TextParseBody(BaseModel):
    text: str

# ---------- Auth ----------
def require_admin(authorization: Optional[str] = Header(None)):
    if not authorization:
        raise HTTPException(401, "Falta token")
    token = authorization.replace("Bearer ", "").strip()
    if token != ADMIN_TOKEN:
        raise HTTPException(401, "Token inválido")
    return True

@api_router.post("/admin/login")
async def admin_login(body: LoginBody):
    if body.username == ADMIN_USER and body.password == ADMIN_PASSWORD:
        return {"token": ADMIN_TOKEN}
    raise HTTPException(401, "Usuario o contraseña incorrectos")

# ---------- Menus ----------
@api_router.get("/menus")
async def list_menus(provider: Optional[str] = None):
    q = {}
    if provider:
        q["provider"] = provider
    docs = await db.menus.find(q, {"_id": 0}).sort("created_at", -1).to_list(200)
    return docs

@api_router.get("/menus/active")
async def get_active_menus():
    # returns latest per provider
    result = {}
    for p in ["mara", "sabrositos"]:
        doc = await db.menus.find_one({"provider": p}, {"_id": 0}, sort=[("created_at", -1)])
        result[p] = doc
    return result

@api_router.post("/menus")
async def create_menu(body: MenuCreate, _=Depends(require_admin)):
    menu = Menu(**body.model_dump())
    await db.menus.insert_one(menu.model_dump())
    return menu

@api_router.put("/menus/{menu_id}")
async def update_menu(menu_id: str, body: MenuCreate, _=Depends(require_admin)):
    upd = body.model_dump()
    await db.menus.update_one({"id": menu_id}, {"$set": upd})
    doc = await db.menus.find_one({"id": menu_id}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "No encontrado")
    return doc

@api_router.delete("/menus/{menu_id}")
async def delete_menu(menu_id: str, _=Depends(require_admin)):
    await db.menus.delete_one({"id": menu_id})
    # cascade selections for this menu
    await db.selections.delete_many({"menu_id": menu_id})
    return {"ok": True}

# ---------- Excel parser (Mara format) ----------
def parse_mara_excel(file_bytes: bytes) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def is_day_header(v):
        if not isinstance(v, str):
            return None
        s = v.strip().lower()
        mapping = {"lunes": "lunes", "martes": "martes", "miercoles": "miercoles",
                   "miércoles": "miercoles", "jueves": "jueves", "viernes": "viernes",
                   "sabado": "sabado", "sábado": "sabado", "lunch sabado": "sabado"}
        return mapping.get(s)

    # Collect day-blocks from both left (cols 0..7) and right (cols 8..) halves
    day_blocks: Dict[str, Dict[str, Any]] = {}
    breakfast_items: List[str] = []

    def scan_block(start_row, day_key, date_val, item_col):
        block = {"main": [], "diet": [], "sides": [], "date": None}
        if isinstance(date_val, datetime):
            block["date"] = date_val.date().isoformat()
        i = start_row + 1
        while i < len(rows):
            row = rows[i]
            first = row[item_col - 1] if item_col > 0 else None
            item = row[item_col]
            # stop if we hit another day header
            if isinstance(item, str) and is_day_header(item):
                break
            if isinstance(first, datetime):
                break
            if item and isinstance(item, str):
                txt = item.strip()
                if txt:
                    marker = str(first or "").lower()
                    up = txt.upper()
                    if "dieta" in marker or "dieta" in txt.lower():
                        block["diet"].append(txt.replace("Dieta", "").strip())
                    elif up.startswith("ENS") or up.startswith("PAN") or up.startswith("MANDIOCA"):
                        block["sides"].append(txt)
                    else:
                        block["main"].append(txt)
            i += 1
        return block

    in_breakfast = False
    for idx, row in enumerate(rows):
        # Left half
        if len(row) > 1:
            left_flag = row[0]
            left_val = row[1]
            if isinstance(left_flag, str) and left_flag.strip().upper().startswith("DESAYUNO"):
                in_breakfast = True
                continue
            day = is_day_header(left_val) if left_val else None
            if day:
                blk = scan_block(idx, day, left_flag if isinstance(left_flag, datetime) else None, 1)
                day_blocks[day] = {**day_blocks.get(day, {}), **blk}
        # Right half (col 8..)
        if len(row) > 9:
            right_flag = row[8]
            right_val = row[9]
            day = is_day_header(right_val) if right_val else None
            if day:
                blk = scan_block(idx, day, right_flag if isinstance(right_flag, datetime) else None, 9)
                day_blocks[day] = {**day_blocks.get(day, {}), **blk}
        # Breakfast collection (column 0 items after DESAYUNO header)
        if in_breakfast and idx > 0:
            v = row[0]
            if isinstance(v, str):
                s = v.strip()
                if s and not s.upper().startswith("DESAYUNO"):
                    breakfast_items.append(s)

    days: List[DayMenu] = []
    for d in DAYS_ORDER:
        blk = day_blocks.get(d, {})
        dm = DayMenu(
            day=d,
            date=blk.get("date"),
            main=blk.get("main", []),
            diet=blk.get("diet", []),
            sides=blk.get("sides", []),
            breakfast=breakfast_items if d != "sabado" else [],
        )
        days.append(dm)

    # Week start/end
    week_start = None
    week_end = None
    dates = [d.date for d in days if d.date]
    if dates:
        week_start = min(dates)
        week_end = max(dates)

    return {"provider": "mara", "week_start": week_start, "week_end": week_end,
            "days": [d.model_dump() for d in days]}


# ---------- Text parser (Sabrositos format) ----------
DAY_KEYWORDS = {
    "lunes": "lunes", "martes": "martes", "miercoles": "miercoles", "miércoles": "miercoles",
    "jueves": "jueves", "viernes": "viernes", "sabado": "sabado", "sábado": "sabado",
}

def parse_sabrositos_text(text: str) -> Dict[str, Any]:
    lines = text.split("\n")
    day_blocks: Dict[str, List[str]] = {}
    current = None
    for raw in lines:
        line = raw.strip()
        if not line:
            continue
        low = line.lower().strip(":*# \t")
        # detect day header (single word or "lunes" contained)
        found = None
        for kw, key in DAY_KEYWORDS.items():
            if low == kw or low.startswith(kw + " ") or low.endswith(" " + kw):
                found = key
                break
        if found:
            current = found
            day_blocks.setdefault(current, [])
            continue
        if current is None:
            continue
        # remove leading emojis/plate icons and bullets
        cleaned = re.sub(r"^[^\w¡¿áéíóúñÁÉÍÓÚÑ]+", "", line).strip()
        if cleaned and not cleaned.lower().startswith("menú"):
            day_blocks[current].append(cleaned)

    days: List[DayMenu] = []
    for d in DAYS_ORDER:
        items = day_blocks.get(d, [])
        days.append(DayMenu(day=d, main=items))
    return {"provider": "sabrositos", "week_start": None, "week_end": None,
            "days": [d.model_dump() for d in days]}


@api_router.post("/menus/parse-excel")
async def parse_excel_endpoint(file: UploadFile = File(...), _=Depends(require_admin)):
    content = await file.read()
    parsed = parse_mara_excel(content)
    return parsed

@api_router.post("/menus/parse-text")
async def parse_text_endpoint(body: TextParseBody, _=Depends(require_admin)):
    parsed = parse_sabrositos_text(body.text)
    return parsed

# ---------- Selections ----------
@api_router.post("/selections")
async def create_selection(body: SelectionCreate):
    if not body.user_name.strip():
        raise HTTPException(400, "Nombre requerido")
    menu = await db.menus.find_one({"id": body.menu_id}, {"_id": 0})
    if not menu:
        raise HTTPException(404, "Menú no encontrado")
    sel = Selection(
        user_name=body.user_name.strip(),
        provider=body.provider,
        menu_id=body.menu_id,
        week_start=menu.get("week_start"),
        week_end=menu.get("week_end"),
        choices=body.choices,
    )
    # Upsert: one selection per (user_name lower, menu_id)
    existing = await db.selections.find_one({
        "menu_id": body.menu_id,
        "user_name_lc": sel.user_name.lower()
    })
    doc = sel.model_dump()
    doc["user_name_lc"] = sel.user_name.lower()
    if existing:
        doc["id"] = existing["id"]
        doc["created_at"] = existing.get("created_at", doc["created_at"])
        await db.selections.update_one({"id": existing["id"]}, {"$set": doc})
    else:
        await db.selections.insert_one(doc)
    return sel

@api_router.get("/selections")
async def list_selections(provider: Optional[str] = None, menu_id: Optional[str] = None,
                          _=Depends(require_admin)):
    q = {}
    if provider: q["provider"] = provider
    if menu_id: q["menu_id"] = menu_id
    docs = await db.selections.find(q, {"_id": 0, "user_name_lc": 0}).sort("created_at", 1).to_list(2000)
    return docs

@api_router.delete("/selections/reset")
async def reset_selections(provider: Optional[str] = None, _=Depends(require_admin)):
    q = {}
    if provider: q["provider"] = provider
    r = await db.selections.delete_many(q)
    return {"deleted": r.deleted_count}

@api_router.get("/selections/export")
async def export_selections(provider: str, menu_id: Optional[str] = None,
                            token: Optional[str] = None):
    # Allow token via query param for direct browser download
    if token != ADMIN_TOKEN:
        raise HTTPException(401, "Token inválido")
    q = {"provider": provider}
    if menu_id:
        q["menu_id"] = menu_id
    else:
        # use latest menu for provider
        menu = await db.menus.find_one({"provider": provider}, {"_id": 0}, sort=[("created_at", -1)])
        if not menu:
            raise HTTPException(404, "No hay menú")
        q["menu_id"] = menu["id"]
        menu_id = menu["id"]

    menu = await db.menus.find_one({"id": q["menu_id"]}, {"_id": 0})
    sels = await db.selections.find(q, {"_id": 0, "user_name_lc": 0}).to_list(2000)

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = workbook.add_worksheet("Pedidos")
    header_fmt = workbook.add_format({"bold": True, "bg_color": "#DC2626", "font_color": "white", "border": 1})
    cell_fmt = workbook.add_format({"border": 1, "text_wrap": True, "valign": "top"})
    name_fmt = workbook.add_format({"border": 1, "bold": True, "valign": "top"})

    # Header: Nombre | for each day: Desayuno / Principal / Dieta / Acompañamiento
    ws.set_column(0, 0, 22)
    headers = ["Nombre"]
    day_col_map = []
    col = 1
    for d in DAYS_ORDER:
        day_data = next((x for x in (menu.get("days") or []) if x["day"] == d), None)
        if not day_data:
            continue
        has_breakfast = bool(day_data.get("breakfast"))
        has_main = bool(day_data.get("main"))
        has_diet = bool(day_data.get("diet"))
        has_sides = bool(day_data.get("sides"))
        date_str = day_data.get("date") or ""
        label = f"{DAY_LABELS[d]}" + (f" ({date_str})" if date_str else "")
        if has_breakfast:
            headers.append(f"{label} - Desayuno"); day_col_map.append((d, "breakfast", col)); col += 1
        if has_main:
            headers.append(f"{label} - Principal"); day_col_map.append((d, "main", col)); col += 1
        if has_diet:
            headers.append(f"{label} - Dieta"); day_col_map.append((d, "diet", col)); col += 1
        if has_sides:
            headers.append(f"{label} - Acompañamiento"); day_col_map.append((d, "side", col)); col += 1

    for i, h in enumerate(headers):
        ws.write(0, i, h, header_fmt)
        ws.set_column(i, i, max(18, min(40, len(h) + 4)))
    ws.freeze_panes(1, 1)

    for r, s in enumerate(sels, start=1):
        ws.write(r, 0, s.get("user_name", ""), name_fmt)
        for (d, kind, cidx) in day_col_map:
            val = ""
            day_choice = (s.get("choices") or {}).get(d) or {}
            if kind == "breakfast":
                val = day_choice.get("breakfast") or ""
            elif kind == "main":
                val = day_choice.get("main") or ""
                # if user picked diet as main, mark
                if not val and day_choice.get("diet"):
                    val = f"[Dieta] {day_choice.get('diet')}"
            elif kind == "diet":
                val = day_choice.get("diet") or ""
            elif kind == "side":
                val = day_choice.get("side") or ""
            ws.write(r, cidx, val, cell_fmt)

    workbook.close()
    output.seek(0)
    filename = f"pedidos_{provider}_{menu.get('week_start') or datetime.now().date().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/")
async def root():
    return {"ok": True, "app": "nasser-menus"}


app.include_router(api_router)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
